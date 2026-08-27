"""Impor playlist dari Excel/CSV + pembuat file template.

Format kolom (nama kolom tidak peka huruf besar/kecil):

    playlist | url | judul

  - `url`      wajib. Link YouTube/TikTok/Facebook/Instagram.
  - `playlist` opsional. Kalau diisi, video masuk ke playlist bernama itu
               (dibuat otomatis kalau belum ada). Kosong = masuk ke playlist
               yang sedang dibuka.
  - `judul`    opsional. Kalau kosong, judul asli bisa diambil lewat tombol Periksa.

Satu file boleh berisi banyak playlist sekaligus — praktis untuk menyiapkan
beberapa ruangan dalam sekali kerja.
"""
import csv
import io

HEADERS = ["playlist", "url", "judul"]

CONTOH = [
    ["Kelas Balita", "https://www.youtube.com/watch?v=dQw4w9WgXcQ", "Lagu pembuka"],
    ["Kelas Balita", "https://www.youtube.com/watch?v=9bZkp7q19f0", ""],
    ["Kelas Batita", "https://www.youtube.com/watch?v=dQw4w9WgXcQ", ""],
]

CATATAN = [
    "Cara pakai:",
    "1. Isi kolom url dengan link video. Ini satu-satunya kolom yang wajib.",
    "2. Kolom playlist boleh dikosongkan - videonya masuk ke playlist yang sedang dibuka.",
    "3. Kalau diisi, playlist dengan nama itu dibuat otomatis bila belum ada.",
    "4. Kolom judul boleh dikosongkan - pakai tombol Periksa di aplikasi untuk",
    "   mengambil judul asli dari platform.",
    "5. Urutan video mengikuti urutan baris.",
    "6. Baris contoh di bawah boleh dihapus atau ditimpa.",
]


def _norm(value) -> str:
    return "" if value is None else str(value).strip()


def build_template_xlsx() -> bytes | None:
    """Template .xlsx. Return None kalau openpyxl tidak tersedia."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Playlist"

    head_fill = PatternFill("solid", fgColor="4F81BD")
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for r, row in enumerate(CONTOH, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=value)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 34
    ws.freeze_panes = "A2"

    notes = wb.create_sheet("Petunjuk")
    notes.column_dimensions["A"].width = 96
    for r, line in enumerate(CATATAN, start=1):
        cell = notes.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv() -> bytes:
    """Cadangan kalau openpyxl tidak ada. BOM supaya Excel membaca UTF-8 dengan benar."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    w.writerows(CONTOH)
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Playlist"] if "Playlist" in wb.sheetnames else wb.worksheets[0]
    return [[_norm(c) for c in row] for row in ws.iter_rows(values_only=True)]


def _rows_from_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", "replace")
    # Excel di sebagian lokal menyimpan CSV dengan titik koma.
    sample = text[:2000]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    return [[_norm(c) for c in row] for row in csv.reader(io.StringIO(text), delimiter=delim)]


def parse(filename: str, data: bytes) -> list[dict]:
    """Baca file jadi daftar baris {playlist, url, judul, baris}.

    Melempar ValueError kalau format tidak dikenali atau kolom url tidak ada.
    """
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        try:
            rows = _rows_from_xlsx(data)
        except ImportError:
            raise ValueError(
                "Dukungan .xlsx belum terpasang. Jalankan ulang aplikasi "
                "(dependency dipasang otomatis), atau simpan file sebagai CSV."
            )
    elif name.endswith((".csv", ".txt")):
        rows = _rows_from_csv(data)
    else:
        raise ValueError("Format tidak dikenali. Pakai .xlsx atau .csv.")

    rows = [r for r in rows if any(r)]
    if not rows:
        raise ValueError("File-nya kosong.")

    header = [h.lower() for h in rows[0]]
    if "url" not in header:
        raise ValueError(
            "Baris pertama harus berisi nama kolom, dan salah satunya `url`. "
            "Download templatenya kalau ragu."
        )
    idx = {key: (header.index(key) if key in header else None) for key in HEADERS}

    def cell(row, key):
        i = idx[key]
        return row[i] if i is not None and i < len(row) else ""

    out = []
    for n, row in enumerate(rows[1:], start=2):
        url = cell(row, "url")
        if not url:
            continue
        out.append({
            "baris": n,
            "url": url,
            "playlist": cell(row, "playlist"),
            "judul": cell(row, "judul"),
        })
    return out
